from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent.parent

SEVERITY_FILLS = {
    "high": PatternFill(fill_type="solid", fgColor="F4CCCC"),
    "medium": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "low": PatternFill(fill_type="solid", fgColor="D9EAD3"),
}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="EEF2F7")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9DEE7"),
    right=Side(style="thin", color="D9DEE7"),
    top=Side(style="thin", color="D9DEE7"),
    bottom=Side(style="thin", color="D9DEE7"),
)


def build_issue_comment(issue):
    return "\n".join(
        [
            f"Severity: {issue['severity']}",
            f"Type: {issue['type']}",
            f"Metric: {issue['metric'] or ''}",
            f"Period: {issue['period'] or ''}",
            f"Message: {issue['message']}",
            f"Business impact: {issue['business_impact']}",
            f"Suggested fix: {issue['suggested_fix']}",
        ]
    )


def add_or_append_comment(cell, comment_text):
    if cell.comment is None:
        cell.comment = Comment(comment_text, "Finance Report Checker")
        return

    existing_text = cell.comment.text
    cell.comment = Comment(
        f"{existing_text}\n\nFinance Report Checker:\n{comment_text}",
        cell.comment.author or "Finance Report Checker",
    )


def style_table_cell(cell, bold=False, fill=None):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if bold:
        cell.font = Font(bold=True)
    if fill is not None:
        cell.fill = fill


def write_table_headers(sheet, row_index, headers):
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=header)
        style_table_cell(cell, bold=True, fill=HEADER_FILL)


def add_qa_report_sheet(workbook, report):
    if "QA_Report" in workbook.sheetnames:
        workbook.remove(workbook["QA_Report"])

    sheet = workbook.create_sheet("QA_Report", 0)
    sheet["A1"] = "Finance Report Checker - QA Report"
    sheet["A1"].font = Font(bold=True, size=20)
    sheet.row_dimensions[1].height = 30

    summary_rows = [
        ("Workbook", report["workbook"]),
        ("Risk Score", report["risk_score"]),
        ("Risk Level", report["risk_level"]),
        ("Executive Summary", report["executive_summary"]),
        ("High Issues", report["issue_counts"]["high"]),
        ("Medium Issues", report["issue_counts"]["medium"]),
        ("Low Issues", report["issue_counts"]["low"]),
    ]

    for row_index, (label, value) in enumerate(summary_rows, start=3):
        label_cell = sheet.cell(row=row_index, column=1, value=label)
        value_cell = sheet.cell(row=row_index, column=2, value=value)
        label_cell.font = Font(bold=True)
        label_cell.fill = SUMMARY_FILL
        label_cell.border = THIN_BORDER
        label_cell.alignment = Alignment(wrap_text=True, vertical="top")
        value_cell.border = THIN_BORDER
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_index].height = 22

    sheet.row_dimensions[6].height = 58

    stats_title_row = 12
    sheet.cell(row=stats_title_row, column=1, value="Sheet Statistics")
    sheet.cell(row=stats_title_row, column=1).font = Font(bold=True, size=13)
    sheet.row_dimensions[stats_title_row].height = 24

    stats_headers = [
        "Sheet",
        "Used Range",
        "Non-empty Cells",
        "Formula Cells",
        "Numeric Constant Cells",
        "Blank Cells Inside Used Range",
    ]
    stats_header_row = stats_title_row + 1
    write_table_headers(sheet, stats_header_row, stats_headers)
    sheet.row_dimensions[stats_header_row].height = 30

    for row_index, sheet_stats in enumerate(report["sheets"], start=stats_header_row + 1):
        values = [
            sheet_stats["name"],
            sheet_stats["used_range"],
            sheet_stats["non_empty_cells"],
            sheet_stats["formula_cells"],
            sheet_stats["numeric_constant_cells"],
            sheet_stats["blank_cells_inside_used_range"],
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            style_table_cell(cell)
        sheet.row_dimensions[row_index].height = 24

    issues_title_row = stats_header_row + max(len(report["sheets"]), 1) + 3
    sheet.cell(row=issues_title_row, column=1, value="Issues")
    sheet.cell(row=issues_title_row, column=1).font = Font(bold=True, size=13)
    sheet.row_dimensions[issues_title_row].height = 24

    headers = [
        "Severity",
        "Type",
        "Sheet",
        "Cell",
        "Metric",
        "Period",
        "Message",
        "Business Impact",
        "Suggested Fix",
    ]
    header_row = issues_title_row + 1
    first_issue_row = header_row + 1
    write_table_headers(sheet, header_row, headers)
    sheet.row_dimensions[header_row].height = 32

    for row_index, issue in enumerate(report["issues"], start=first_issue_row):
        values = [
            issue["severity"],
            issue["type"],
            issue["sheet"],
            issue["cell"],
            issue["metric"],
            issue["period"],
            issue["message"],
            issue["business_impact"],
            issue["suggested_fix"],
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            style_table_cell(cell)

        severity_fill = SEVERITY_FILLS.get(issue["severity"])
        if severity_fill is not None:
            sheet.cell(row=row_index, column=1).fill = severity_fill
        sheet.row_dimensions[row_index].height = 58

    sheet.freeze_panes = f"A{first_issue_row}"
    last_issue_row = max(first_issue_row, first_issue_row + len(report["issues"]) - 1)
    sheet.auto_filter.ref = f"A{header_row}:I{last_issue_row}"
    auto_width_columns(sheet)
    set_qa_report_column_widths(sheet)


def auto_width_columns(sheet):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)


def set_qa_report_column_widths(sheet):
    widths = {
        "A": 18,
        "B": 36,
        "C": 18,
        "D": 12,
        "E": 22,
        "F": 16,
        "G": 46,
        "H": 52,
        "I": 56,
    }

    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width


def save_annotated_workbook(workbook_path, report):
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, data_only=False)

    if "QA_Report" in workbook.sheetnames:
        workbook.remove(workbook["QA_Report"])

    for issue in report["issues"]:
        if issue["sheet"] not in workbook.sheetnames:
            continue

        sheet = workbook[issue["sheet"]]
        cell = sheet[issue["cell"]]
        fill = SEVERITY_FILLS.get(issue["severity"])
        if fill is not None:
            cell.fill = fill
        add_or_append_comment(cell, build_issue_comment(issue))

    add_qa_report_sheet(workbook, report)

    checked_path = PROJECT_ROOT / f"checked_{workbook_path.name}"
    workbook.save(checked_path)
    return checked_path
