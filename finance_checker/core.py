from pathlib import Path

from openpyxl import load_workbook

from finance_checker.rules import (
    cell_address,
    detect_row_issues,
    is_empty,
    is_formula,
    is_numeric_constant,
    prioritize_issues,
    should_ignore_issue_sheet,
)
from finance_checker.scoring import add_risk_summary


def get_used_bounds(sheet):
    """Return the bounding box of non-empty cells, or None for an empty sheet."""
    min_row = min_col = None
    max_row = max_col = None

    for row in sheet.iter_rows():
        for cell in row:
            if is_empty(cell.value):
                continue

            min_row = cell.row if min_row is None else min(min_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_row = cell.row if max_row is None else max(max_row, cell.row)
            max_col = cell.column if max_col is None else max(max_col, cell.column)

    if min_row is None:
        return None

    return trim_at_blank_separators(sheet, min_row, min_col, max_row, max_col)


def trim_at_blank_separators(sheet, min_row, min_col, max_row, max_col):
    """Ignore notes or metadata separated from the main report by a blank row/column."""
    for column in range(min_col, max_col + 1):
        is_blank_separator = all(
            is_empty(sheet.cell(row=row, column=column).value)
            for row in range(min_row, max_row + 1)
        )
        if is_blank_separator:
            max_col = column - 1
            break

    for row in range(min_row, max_row + 1):
        is_blank_separator = all(
            is_empty(sheet.cell(row=row, column=column).value)
            for column in range(min_col, max_col + 1)
        )
        if is_blank_separator:
            max_row = row - 1
            break

    return min_row, min_col, max_row, max_col


def summarize_sheet(sheet, bounds):
    if bounds is None:
        return {
            "name": sheet.title,
            "used_range": "",
            "non_empty_cells": 0,
            "formula_cells": 0,
            "numeric_constant_cells": 0,
            "blank_cells_inside_used_range": 0,
        }

    min_row, min_col, max_row, max_col = bounds
    non_empty_cells = 0
    formula_cells = 0
    numeric_constant_cells = 0

    for row in sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            if is_empty(cell.value):
                continue

            non_empty_cells += 1
            if is_formula(cell):
                formula_cells += 1
            elif is_numeric_constant(cell):
                numeric_constant_cells += 1

    total_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
    used_range = f"{cell_address(min_row, min_col)}:{cell_address(max_row, max_col)}"

    return {
        "name": sheet.title,
        "used_range": used_range,
        "non_empty_cells": non_empty_cells,
        "formula_cells": formula_cells,
        "numeric_constant_cells": numeric_constant_cells,
        "blank_cells_inside_used_range": total_cells - non_empty_cells,
    }


def analyze_workbook(file_path: str) -> dict:
    workbook_path = Path(file_path)
    workbook = load_workbook(workbook_path, data_only=False)
    report = {
        "workbook": workbook_path.name,
        "sheets": [],
        "issues": [],
    }

    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible":
            continue

        bounds = get_used_bounds(sheet)
        report["sheets"].append(summarize_sheet(sheet, bounds))
        if not should_ignore_issue_sheet(sheet):
            report["issues"].extend(detect_row_issues(sheet, bounds))

    report["issues"] = prioritize_issues(report["issues"])
    add_risk_summary(report)
    return report
