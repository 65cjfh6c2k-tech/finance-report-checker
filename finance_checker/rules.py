import re
from collections import Counter

from openpyxl.utils import get_column_letter, column_index_from_string


CELL_REFERENCE_RE = re.compile(
    r"(?<![A-Za-z0-9_])(\$?[A-Z]{1,3}\$?\d+)(?![A-Za-z0-9_])"
)

IGNORED_ISSUE_SHEETS = {"sources", "notes", "assumptions", "instructions"}
AGGREGATE_PERIOD_HEADERS = {"fy total", "total", "ytd", "ltm"}
VOLATILE_VARIANCE_METRICS = [
    "other income",
    "other expense",
    "other income/(expense)",
    "fx",
    "foreign exchange",
]
NOISY_TOTAL_METRICS = {"total net sales", "total cost of sales"}


def is_empty(value):
    """Treat None and empty strings as blank spreadsheet cells."""
    return value is None or value == ""


def is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")


def is_numeric_constant(cell):
    return isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)


def cell_address(row, column):
    return f"{get_column_letter(column)}{row}"


def get_metric_for_row(sheet, row_number):
    """Use column A as the finance metric label for an issue row."""
    value = sheet.cell(row=row_number, column=1).value
    if is_empty(value):
        return None
    return str(value)


def get_period_for_column(sheet, column_number):
    """Use row 1 as the reporting period/header label for an issue column."""
    value = sheet.cell(row=1, column=column_number).value
    if is_empty(value):
        return None
    return str(value)


def normalized_label(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def should_ignore_issue_sheet(sheet):
    return normalized_label(sheet.title) in IGNORED_ISSUE_SHEETS


def is_aggregate_period(period):
    return normalized_label(period) in AGGREGATE_PERIOD_HEADERS


def is_volatile_variance_metric(metric):
    normalized_metric = normalized_label(metric)
    return any(term in normalized_metric for term in VOLATILE_VARIANCE_METRICS)


def is_total_or_subtotal_metric(metric):
    normalized_metric = normalized_label(metric)
    return "total" in normalized_metric or "subtotal" in normalized_metric


def should_skip_total_formula_check(metric):
    return normalized_label(metric) in NOISY_TOTAL_METRICS


def get_business_impact(metric):
    if metric is None:
        return "This may affect the reliability of the financial report."

    if "Revenue" in metric:
        return "This may affect revenue analysis and top-line reporting."
    if "COGS" in metric:
        return "This may affect cost of goods sold analysis and gross margin calculations."
    if "Gross Profit" in metric:
        return "This may affect gross profit and margin analysis."
    if "OPEX" in metric:
        return "This may affect operating expense analysis."
    if "EBITDA" in metric:
        return "This may affect profitability analysis and management reporting."

    return "This may affect the reliability of the financial report."


def parse_cell_reference(reference):
    match = re.fullmatch(r"(\$?)([A-Z]{1,3})(\$?)(\d+)", reference)
    if not match:
        return None

    col_absolute, col_letters, row_absolute, row_text = match.groups()
    return {
        "col_absolute": bool(col_absolute),
        "col": column_index_from_string(col_letters),
        "row_absolute": bool(row_absolute),
        "row": int(row_text),
    }


def formula_pattern(formula, origin_row, origin_column):
    """Normalize formulas so copied formulas have the same relative pattern."""

    def replace_reference(match):
        reference = match.group(1)
        parsed = parse_cell_reference(reference)
        if parsed is None:
            return reference

        if parsed["col_absolute"]:
            col_part = f"C${parsed['col']}"
        else:
            col_part = f"C{parsed['col'] - origin_column:+d}"

        if parsed["row_absolute"]:
            row_part = f"R${parsed['row']}"
        else:
            row_part = f"R{parsed['row'] - origin_row:+d}"

        return f"[{col_part},{row_part}]"

    return CELL_REFERENCE_RE.sub(replace_reference, formula.upper())


def add_issue(issues, seen, severity, issue_type, sheet, cell, message, suggested_fix):
    key = (issue_type, sheet.title, cell.coordinate)
    if key in seen:
        return

    metric = get_metric_for_row(sheet, cell.row)
    period = get_period_for_column(sheet, cell.column)

    seen.add(key)
    issues.append(
        {
            "severity": severity,
            "type": issue_type,
            "sheet": sheet.title,
            "cell": cell.coordinate,
            "metric": metric,
            "period": period,
            "message": message,
            "suggested_fix": suggested_fix,
            "business_impact": get_business_impact(metric),
        }
    )


def detect_row_issues(sheet, bounds):
    if bounds is None:
        return []

    issues = []
    seen = set()
    min_row, min_col, max_row, max_col = bounds

    for row_number in range(min_row, max_row + 1):
        row_cells = [
            sheet.cell(row=row_number, column=column)
            for column in range(min_col, max_col + 1)
        ]

        non_empty = [cell for cell in row_cells if not is_empty(cell.value)]
        formulas = [cell for cell in non_empty if is_formula(cell)]
        numeric_constants = [cell for cell in non_empty if is_numeric_constant(cell)]

        # A numeric value surrounded by copied formulas is often an accidental override.
        value_cells = formulas + numeric_constants
        if len(formulas) >= 2 and len(formulas) >= max(1, len(value_cells) / 2):
            formula_columns = {cell.column for cell in formulas}
            for cell in numeric_constants:
                has_formula_left = any(column < cell.column for column in formula_columns)
                has_formula_right = any(column > cell.column for column in formula_columns)
                if has_formula_left and has_formula_right:
                    add_issue(
                        issues,
                        seen,
                        "medium",
                        "hardcoded_value_among_formulas",
                        sheet,
                        cell,
                        "Numeric constant found in a row that mostly contains formulas.",
                        "Check whether this value should be a formula copied from adjacent periods.",
                    )

        # A blank cell between filled cells can indicate a missing value or formula.
        for index, cell in enumerate(row_cells):
            if not is_empty(cell.value):
                continue

            has_left_value = any(not is_empty(left.value) for left in row_cells[:index])
            has_right_value = any(not is_empty(right.value) for right in row_cells[index + 1 :])
            if has_left_value and has_right_value:
                add_issue(
                    issues,
                    seen,
                    "low",
                    "blank_cell_between_values",
                    sheet,
                    cell,
                    "Blank cell found between non-empty cells in the same row.",
                    "Check whether this blank cell should contain a value or formula.",
                )

        detect_formula_inconsistencies(sheet, formulas, issues, seen)
        detect_variance_spikes(sheet, row_cells, issues, seen)
        detect_total_formula_mismatches(sheet, row_cells, issues, seen)

    return issues


def detect_total_formula_mismatches(sheet, row_cells, issues, seen):
    metric = get_metric_for_row(sheet, row_cells[0].row)
    if (
        metric is None
        or not is_total_or_subtotal_metric(metric)
        or should_skip_total_formula_check(metric)
    ):
        return

    expected_rows = get_expected_detail_rows(sheet, row_cells[0].row)
    if not expected_rows:
        return

    for cell in row_cells:
        if cell.column == 1:
            continue
        if not is_formula(cell):
            continue
        if is_aggregate_period(get_period_for_column(sheet, cell.column)):
            continue
        if not formula_includes_expected_rows(cell.value, cell.column, expected_rows):
            add_issue(
                issues,
                seen,
                "high",
                "total_formula_mismatch",
                sheet,
                cell,
                "Total formula may not include all expected detail rows.",
                "Review whether the total formula includes all relevant detail rows for this period.",
            )


def get_expected_detail_rows(sheet, total_row):
    expected_rows = []

    for row_number in range(total_row - 1, 1, -1):
        metric = get_metric_for_row(sheet, row_number)
        if metric is None or is_total_or_subtotal_metric(metric):
            break
        if row_has_formula_values(sheet, row_number):
            break
        expected_rows.append(row_number)

    expected_rows.reverse()
    return expected_rows


def row_has_formula_values(sheet, row_number):
    for column_number in range(2, sheet.max_column + 1):
        if is_aggregate_period(get_period_for_column(sheet, column_number)):
            continue
        if is_formula(sheet.cell(row=row_number, column=column_number)):
            return True
    return False


def formula_includes_expected_rows(formula, column_number, expected_rows):
    column_letter = get_column_letter(column_number)
    covered_rows = referenced_rows_for_column(formula, column_letter)
    return set(expected_rows).issubset(covered_rows)


def referenced_rows_for_column(formula, column_letter):
    formula_text = str(formula).upper()
    escaped_column = re.escape(column_letter.upper())
    covered_rows = set()

    range_pattern = re.compile(
        rf"(?:'[^']+'!)?\$?{escaped_column}\$?(\d+)\s*:\s*"
        rf"(?:'[^']+'!)?\$?{escaped_column}\$?(\d+)"
    )
    for start_text, end_text in range_pattern.findall(formula_text):
        start_row = int(start_text)
        end_row = int(end_text)
        low_row, high_row = sorted((start_row, end_row))
        covered_rows.update(range(low_row, high_row + 1))

    cell_pattern = re.compile(
        rf"(?<![A-Z0-9_])(?:'[^']+'!)?\$?{escaped_column}\$?(\d+)(?![A-Z0-9_])"
    )
    covered_rows.update(int(row_text) for row_text in cell_pattern.findall(formula_text))
    return covered_rows


def detect_variance_spikes(sheet, row_cells, issues, seen):
    metric = get_metric_for_row(sheet, row_cells[0].row)
    if metric is None or is_volatile_variance_metric(metric):
        return

    # Compare adjacent period columns, starting after the metric label column.
    for previous_cell, current_cell in zip(row_cells, row_cells[1:]):
        if previous_cell.column == 1:
            continue
        if is_aggregate_period(get_period_for_column(sheet, previous_cell.column)):
            continue
        if is_aggregate_period(get_period_for_column(sheet, current_cell.column)):
            continue
        if not is_numeric_constant(previous_cell) or not is_numeric_constant(current_cell):
            continue
        if previous_cell.value == 0:
            continue

        percentage_change = abs(current_cell.value - previous_cell.value) / abs(
            previous_cell.value
        )
        if percentage_change > 0.5:
            add_issue(
                issues,
                seen,
                "medium",
                "variance_spike_between_periods",
                sheet,
                current_cell,
                "Value changed by more than 50% compared with the previous period.",
                "Review whether this period-over-period movement is expected and properly explained.",
            )


def detect_formula_inconsistencies(sheet, formulas, issues, seen):
    if len(formulas) < 2:
        return

    patterns = {
        cell.coordinate: formula_pattern(cell.value, cell.row, cell.column)
        for cell in formulas
    }
    pattern_counts = Counter(patterns.values())
    majority_pattern, majority_count = pattern_counts.most_common(1)[0]

    # In rows with at least three formulas, flag the formula that differs from the row's
    # dominant relative pattern. For two neighboring formulas, flag the right-hand cell.
    if len(formulas) >= 3 and majority_count >= 2:
        for cell in formulas:
            if patterns[cell.coordinate] != majority_pattern:
                add_formula_issue(sheet, cell, issues, seen)
        return

    formulas_by_column = sorted(formulas, key=lambda cell: cell.column)
    for left, right in zip(formulas_by_column, formulas_by_column[1:]):
        if right.column != left.column + 1:
            continue
        if patterns[left.coordinate] != patterns[right.coordinate]:
            add_formula_issue(sheet, right, issues, seen)


def add_formula_issue(sheet, cell, issues, seen):
    if is_aggregate_period(get_period_for_column(sheet, cell.column)):
        return

    add_issue(
        issues,
        seen,
        "medium",
        "formula_inconsistency_in_row",
        sheet,
        cell,
        "Formula pattern differs from neighboring formulas in the same row.",
        "Review whether this formula should match the copied pattern used nearby.",
    )


def prioritize_issues(issues):
    total_mismatch_cells = {
        (issue["sheet"], issue["cell"])
        for issue in issues
        if issue["type"] == "total_formula_mismatch"
    }

    prioritized = []
    for issue in issues:
        same_cell = (issue["sheet"], issue["cell"])
        if (
            issue["type"] == "formula_inconsistency_in_row"
            and same_cell in total_mismatch_cells
        ):
            continue
        prioritized.append(issue)

    return prioritized
